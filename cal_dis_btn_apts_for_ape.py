import csv
import pandas as pd
from geographiclib.geodesic import Geodesic
import time
import os
from openpyxl import load_workbook
import json


def cal_dis_btn_apts_for_ap(runway: int, airplaneType: str, range: int):
    url = r"D:/NiceFlight/Airline_Manager/flight_plan.xlsx"
    df = pd.read_excel(url, header=0, sheet_name="airportsData")
    json_data = df.to_json(orient="records", force_ascii=False, indent=4)
    json_data = json.loads(json_data)

    selectAirports = []
    for _ in json_data:
        if int(_["Max Runway(ft)"]) >= runway:
            selectAirports.append(_)

    _startAirport = ""
    _endAirport = ""
    dis = 0
    output_list = []
    f_list = []

    for i in selectAirports:
        for j in selectAirports:
            if i["Id"] == j["Id"]:
                continue
            startAirport = i["Name"].strip()
            startIATA = i["IATA Code"].strip()
            start_lat = float(i["Lat"])
            start_lng = float(i["Lng"])

            endAirport = j["Name"].strip()
            endIATA = j["IATA Code"].strip()
            end_lat = float(j["Lat"])
            end_lng = float(j["Lng"])
            # route = sorted([startIATA, endIATA])

            # routes.append(route)
            # print(f"{startAirport}, {startIATA}({start_lat}, {start_lng})-{endAirport}, {endIATA}({end_lat}, {end_lng})")

            # model: sphere
            geod = Geodesic(6371000, 0)
            line = geod.InverseLine(start_lat, start_lng, end_lat, end_lng)
            distance = line.s13 / 1000  # Convert to kilometers

            # calculate the limit range of the plane
            if 18500 < distance:
                print(f"{startAirport}-{endAirport} - Distance: {int(distance)} km")
                output_list.append([sorted([startIATA, endIATA]), int(distance), sorted([startAirport, endAirport])])

                for _ in output_list:
                    if _ not in f_list:
                        f_list.append(_)

            # find the Max range of the plane
            if distance > dis:
                dis = distance
                _startAirport = startAirport
                _endAirport = endAirport

    print(f"The most far airports for {airplaneType} are {_startAirport} and {_endAirport}.\nThe distance is {dis} km!")

    # sortd data
    f_list = sorted(f_list, key=lambda x: x[1], reverse=True)

    with open(f"csv_files/The_distance_of_two_airports_for_{airplaneType}.csv", "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerows(f_list)


def my_hubs_to_every_apts_distance(airplaneType: str, ReqRunway: int, range: int):

    startList = ["AQP", "TPE", "PER", "CHC", "HNL", "VIE", "AUH", "ORD", "SCL", "JNB", "KEF", "SIN"]

    """Find my hub's to every airports distance"""
    url = r"D:/NiceFlight/Airline_Manager/flight_plan.xlsx"
    airportsDF = pd.read_excel(url, header=0, sheet_name="airportsData")

    output_list1 = []
    output_list2 = []
    for _ in startList:
        start_airport = airportsDF[airportsDF["IATA Code"].str.strip() == _]
        # print(start_airport)
        if int(start_airport["Max Runway(ft)"].values[0]) >= ReqRunway:
            start_IATA = start_airport["IATA Code"].str.strip().values[0]
            start_Lat = start_airport["Lat"].values[0]
            start_Lng = start_airport["Lng"].values[0]
            # print(start_IATA, start_Lat, start_Lng)

        selectAirports = []
        for airport in airportsDF.iloc:  # df to list
            if int(airport["Max Runway(ft)"]) >= ReqRunway and airport["IATA Code"] != _:
                selectAirports.append(airport.tolist())
        # print(selectAirports)

        for i in selectAirports:
            end_IATA = i[5].strip()
            end_Lat = i[7]
            end_Lng = i[8]
            if start_IATA == end_IATA:
                continue

            # model: sphere
            geod = Geodesic(6371000, 0)
            line = geod.InverseLine(start_Lat, start_Lng, end_Lat, end_Lng)
            distance = line.s13 / 1000  # Convert to kilometers

            if float(range * 0.5) <= distance <= float(range):
                print(f"{start_IATA}-{end_IATA} - Distance: {round(distance)} km")
                output_list1.append((f"{start_IATA}-{end_IATA}", round(distance)))

            output_list1 = sorted(output_list1, key=lambda x: x[1], reverse=True)  # sorted output_list1

            if 18500 <= distance:
                print(f"{start_IATA}-{end_IATA} - Distance: {round(distance)} km")
                output_list2.append((f"{start_IATA}-{end_IATA}", round(distance)))

            output_list2 = sorted(output_list2, key=lambda x: x[1], reverse=True)  # sorted output_list2

    with open(os.path.join("csv_files", f"The_distance_to_suitable_airports_for_{airplaneType}_{range}.csv"), "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerows(output_list1)

    wb = load_workbook(os.path.join("D:\\", "NiceFlight", "Airline_Manager", "flight_plan.xlsx"))

    sheet_range_times_1 = wb.create_sheet(f"{airplaneType}_{range}")
    sheet_range_times_2 = wb.create_sheet(f"{airplaneType}_{range*2}")
    headers = ["Route", "Distance"]

    for col_idx, header in enumerate(headers, start=1):
        sheet_range_times_1.cell(row=1, column=col_idx, value=header)

    for row_idx, row_data in enumerate(output_list1, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            sheet_range_times_1.cell(row=row_idx, column=col_idx, value=value)

    for col_idx, header in enumerate(headers, start=1):
        sheet_range_times_2.cell(row=1, column=col_idx, value=header)

    for row_idx, row_data in enumerate(output_list2, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            sheet_range_times_2.cell(row=row_idx, column=col_idx, value=value)

    wb.save(os.path.join("D:\\", "NiceFlight", "Airline_Manager", "flight_plan.xlsx"))

    with open(
        os.path.join("csv_files", f"The_distance_to_suitable_airports_for_{airplaneType}_{range*2}.csv"), "w", encoding="utf-8", newline=""
    ) as f:
        writer = csv.writer(f)
        writer.writerows(output_list2)


if __name__ == "__main__":
    # cal_dis_btn_apts_for_ap(11800, "Concorde", 7500)
    # cal_dis_btn_apts_for_ap(9680, "A380", 14500)
    my_hubs_to_every_apts_distance("A380", 9680, 14500)
